import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def convert_coordinates(coord_str):
    """Конвертирует координаты из формата +0691949 в 69-19-49"""
    if not coord_str:
        return ""

    coord_str = coord_str.strip('+')
    if len(coord_str) == 7:
        return f"{coord_str[0:2]}-{coord_str[2:4]}-{coord_str[4:7]}"
    elif len(coord_str) == 6:
        return f"{coord_str[0:2]}-{coord_str[2:4]}-{coord_str[4:6]}"
    return coord_str


def parse_head_section(content):
    """Извлекает данные из секции HEAD"""
    head_data = {}
    head_match = re.search(r'<HEAD>(.*?)</HEAD>', content, re.DOTALL)

    if head_match:
        head_content = head_match.group(1)

        # Извлекаем t_adm (страна)
        adm_match = re.search(r't_adm\s*=\s*(.+)', head_content)
        if adm_match:
            head_data['t_adm'] = adm_match.group(1).strip()

        # Извлекаем t_d_sent (дата отправки)
        sent_match = re.search(r't_d_sent\s*=\s*(.+)', head_content)
        if sent_match:
            head_data['t_d_sent'] = sent_match.group(1).strip()

    return head_data


def parse_notice_block(notice_text):
    """Парсит один блок NOTICE и извлекает данные"""
    data = {}

    # Извлекаем основные параметры
    patterns = {
        't_site_name': r't_site_name\s*=\s*(.+)',
        't_freq_assgn': r't_freq_assgn\s*=\s*(.+)',
        't_long': r't_long\s*=\s*(.+)',
        't_lat': r't_lat\s*=\s*(.+)',
        't_bdwdth_cde': r't_bdwdth_cde\s*=\s*(.+)',
        't_d_adm_ntc': r't_d_adm_ntc\s*=\s*(.+)',
        't_adm_ref_id': r't_adm_ref_id\s*=\s*(.+)',
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, notice_text)
        if match:
            data[key] = match.group(1).strip()

    # Извлекаем данные антенн
    antenna_blocks = re.findall(r'<ANTENNA>(.*?)</ANTENNA>', notice_text, re.DOTALL)

    for antenna in antenna_blocks:
        gain_match = re.search(r't_gain_max\s*=\s*(.+)', antenna)
        height_match = re.search(r't_hgt_agl\s*=\s*(.+)', antenna)
        power_match = re.search(r't_pwr_dbw\s*=\s*(.+)', antenna)

        if gain_match:
            data['t_gain_max'] = gain_match.group(1).strip()
        if height_match:
            data['t_hgt_agl'] = height_match.group(1).strip()
        if power_match:
            data['t_pwr_dbw'] = power_match.group(1).strip()

        # Извлекаем данные принимающей станции
        rx_match = re.search(r'<RX_STATION>(.*?)</RX_STATION>', antenna, re.DOTALL)
        if rx_match:
            rx_content = rx_match.group(1)
            rx_site_match = re.search(r't_site_name\s*=\s*(.+)', rx_content)
            if rx_site_match:
                data['rx_site_name'] = rx_site_match.group(1).strip()

    return data


def parse_txt_file(file_path):
    """Парсит txt файл и возвращает список данных всех станций"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Извлекаем данные из HEAD
    head_data = parse_head_section(content)

    # Разделяем на блоки NOTICE
    notice_blocks = re.findall(r'<NOTICE>(.*?)</NOTICE>', content, re.DOTALL)

    stations_data = []
    for notice in notice_blocks:
        data = parse_notice_block(notice)
        # Добавляем данные из HEAD
        data['t_adm'] = head_data.get('t_adm', '')
        data['t_d_sent'] = head_data.get('t_d_sent', '')
        stations_data.append(data)

    return stations_data, head_data


def link_stations(stations_data):
    """Связывает станции и определяет частоты приёма"""
    # Создаём словарь для быстрого поиска станций по имени
    station_map = {}
    for station in stations_data:
        site_name = station.get('t_site_name', '')
        station_map[site_name] = station

    # Для каждой станции находим частоту приёма
    for station in stations_data:
        rx_site_name = station.get('rx_site_name', '')
        if rx_site_name and rx_site_name in station_map:
            # Частота приёма = частота передачи парной станции
            rx_station = station_map[rx_site_name]
            station['freq_rx'] = rx_station.get('t_freq_assgn', '')
        else:
            station['freq_rx'] = ''

    return stations_data


def determine_sheet_from_adm(adm_code):
    """Определяет лист Excel по коду администрации из HEAD"""
    adm_upper = adm_code.upper()

    if 'KAZ' in adm_upper:
        return 'КАЗ'
    elif 'KGZ' in adm_upper:
        return 'КГЗ'
    elif 'TJK' in adm_upper or 'TAJ' in adm_upper:
        return 'ТЖК'
    elif 'TKM' in adm_upper or 'TUR' in adm_upper:
        return 'ТКМ'
    else:
        return 'КАЗ'  # По умолчанию


def create_sheet_with_data(ws, all_data):
    """Создает лист с данными для ВХОДЯЩИЕ РРЛ"""

    # Заголовок - первая строка (объединенная)
    ws.merge_cells('A1:Q1')
    header_cell = ws['A1']
    header_cell.value = "Учёт статистических данных по частотоприрсвоениям направленных на координацию с АС РУз (ВХОДЯЩИЕ)-РРЛ"
    header_cell.font = Font(bold=True, size=11)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Вторая строка - объединенные заголовки
    ws.merge_cells('A2:B2')
    ws['A2'].value = "Частота, МГц"

    ws.merge_cells('C2:D2')
    ws['C2'].value = "Координаты"

    ws.merge_cells('J2:K2')
    ws['J2'].value = "№ и дата входящего письма"

    ws.merge_cells('L2:M2')
    ws['L2'].value = "№ и дата исходящего письма"

    # Третья строка - подзаголовки
    headers_row3 = [
        "передача", "приём",
        "долгота", "широта",
        "Пункт установки",
        "Ширина\nполосы,\nМГц",
        "Коэф-т\nусиления,\nдБ",
        "Мощность\nпередатчика,\nдБВт",
        "Высота\nантенны, м",
        "первичное", "повторное",
        "первичное", "повторное",
        "Результат согласования\n(согласовано/\nне согласовано)",
        "Примечание",
        "Исполнитель",
        "t_adm_ref_id"
    ]

    # Стиль заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем стили к строке 2
    for col in range(1, 18):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    # Записываем подзаголовки в третью строку
    for col, header in enumerate(headers_row3, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    # Записываем данные начиная с 4-й строки
    row = 4
    for data in all_data:
        # Координаты
        long_coord = convert_coordinates(data.get('t_long', ''))
        lat_coord = convert_coordinates(data.get('t_lat', ''))

        ws.cell(row, 1).value = data.get('t_freq_assgn', '')  # Частота передача
        ws.cell(row, 2).value = data.get('freq_rx', '')  # Частота приём
        ws.cell(row, 3).value = long_coord  # Долгота
        ws.cell(row, 4).value = lat_coord  # Широта
        ws.cell(row, 5).value = data.get('t_site_name', '')  # Пункт установки
        ws.cell(row, 6).value = data.get('t_bdwdth_cde', '')  # Ширина полосы
        ws.cell(row, 7).value = data.get('t_gain_max', '')  # Коэф усиления
        ws.cell(row, 8).value = data.get('t_pwr_dbw', '')  # Мощность
        ws.cell(row, 9).value = data.get('t_hgt_agl', '')  # Высота

        # Формируем номер входящего: t_d_sent + t_d_adm_ntc
        incoming_number = ""
        d_sent = data.get('t_d_sent', '')
        d_adm_ntc = data.get('t_d_adm_ntc', '')
        if d_sent and d_adm_ntc:
            incoming_number = f"{d_sent}/{d_adm_ntc}"
        elif d_sent:
            incoming_number = d_sent
        elif d_adm_ntc:
            incoming_number = d_adm_ntc

        ws.cell(row, 10).value = incoming_number  # № входящего первичное
        ws.cell(row, 11).value = ""  # № входящего повторное
        ws.cell(row, 12).value = ""  # № исходящего первичное
        ws.cell(row, 13).value = ""  # № исходящего повторное
        ws.cell(row, 14).value = ""  # Результат согласования
        ws.cell(row, 15).value = ""  # Примечание
        ws.cell(row, 16).value = ""  # Исполнитель
        ws.cell(row, 17).value = data.get('t_adm_ref_id', '')  # t_adm_ref_id

        # Применяем стиль к ячейкам
        for col in range(1, 18):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9)

        row += 1

    # Настраиваем ширину столбцов
    column_widths = [12, 12, 10, 10, 20, 10, 10, 12, 10, 15, 15, 15, 15, 15, 15, 15, 20]
    for col, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

    # Высота строк
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40


def create_excel(data_by_sheet, output_file):
    """Создает Excel файл с несколькими листами"""
    wb = Workbook()

    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Создаем листы в определенном порядке
    sheet_names = ["КГЗ", "ТЖК", "КАЗ", "ТКМ"]

    for sheet_name in sheet_names:
        ws = wb.create_sheet(sheet_name)

        # Получаем данные для этого листа
        sheet_data = data_by_sheet.get(sheet_name, [])

        # Создаем лист с данными
        create_sheet_with_data(ws, sheet_data)

    wb.save(output_file)
    print(f"✓ Excel файл создан: {output_file}")


def main():
    """Основная функция"""
    # Папка с txt файлами
    input_folder = input("Введите путь к папке с .txt файлами (ВХОДЯЩИЕ РРЛ): ").strip()

    if not os.path.exists(input_folder):
        print("❌ Папка не найдена!")
        return

    # Находим все txt файлы
    txt_files = [f for f in os.listdir(input_folder) if f.endswith('.txt')]

    if not txt_files:
        print("❌ В папке нет .txt файлов!")
        return

    print(f"Найдено {len(txt_files)} файлов\n")

    # Словарь для хранения данных по листам
    data_by_sheet = {
        'КГЗ': [],
        'ТЖК': [],
        'КАЗ': [],
        'ТКМ': []
    }

    # Обрабатываем файлы
    total_stations = 0
    for txt_file in txt_files:
        file_path = os.path.join(input_folder, txt_file)
        print(f"Обработка: {txt_file}...")

        stations_data, head_data = parse_txt_file(file_path)

        # Связываем станции и определяем частоты приёма
        stations_data = link_stations(stations_data)

        target_adm = head_data.get('t_adm', '')

        # Определяем целевой лист по t_adm из HEAD
        target_sheet = determine_sheet_from_adm(target_adm) if target_adm else 'КАЗ'

        # Добавляем данные на соответствующий лист
        data_by_sheet[target_sheet].extend(stations_data)

        total_stations += len(stations_data)
        print(f"  └─ Извлечено {len(stations_data)} станций от {target_adm} → лист '{target_sheet}'\n")

    print(f"📊 Всего станций: {total_stations}")
    print("\n📋 Распределение по листам:")
    for sheet_name, data in data_by_sheet.items():
        if data:
            print(f"  • {sheet_name}: {len(data)} станций")
        else:
            print(f"  • {sheet_name}: 0 станций (пустой)")

    # Создаем Excel файл с уникальным именем
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(input_folder, f"ВХОДЯЩИЕ_РРЛ_{timestamp}.xlsx")
    create_excel(data_by_sheet, output_file)

    print(f"\n✅ Готово! Данные сохранены в: {output_file}")


if __name__ == "__main__":
    main()