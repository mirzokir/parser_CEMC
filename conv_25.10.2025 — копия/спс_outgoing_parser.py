import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def convert_coordinates(coord_str):
    """Конвертирует координаты из формата +0691949 в 69-19-49"""
    if not coord_str:
        return ""

    coord_str = coord_str.strip('+')
    if len(coord_str) == 7:
        return f"{coord_str[0:2]}{coord_str[2:4]}{coord_str[4:7]}"
    elif len(coord_str) == 6:
        return f"{coord_str[0:2]}{coord_str[2:4]}{coord_str[4:6]}"
    return coord_str


def convert_date(date_str):
    """Конвертирует дату из формата 2025-08-13 в 13.08.2025"""
    if not date_str:
        return ""
    try:
        date_obj = datetime.strptime(date_str.strip(), "%Y-%m-%d")
        return date_obj.strftime("%d.%m.%Y")
    except:
        return date_str


def parse_notice_block(notice_text):
    """Парсит один блок NOTICE и извлекает данные"""
    data = {}

    # Извлекаем основные параметры
    patterns = {
        't_site_name': r't_site_name\s*=\s*(.+)',
        't_freq_assgn': r't_freq_assgn\s*=\s*(.+)',
        't_long': r't_long\s*=\s*(.+)',
        't_lat': r't_lat\s*=\s*(.+)',
        't_bdwdth_cde': r't_bdwdth_cde\s*=\s*(.+)',
        't_adm_ref_id': r't_adm_ref_id\s*=\s*(.+)',
        't_d_adm_ntc': r't_d_adm_ntc\s*=\s*(.+)',
        't_d_inuse': r't_d_inuse\s*=\s*(.+)',
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, notice_text)
        if match:
            data[key] = match.group(1).strip()

    # Извлекаем данные антенн
    antenna_blocks = re.findall(r'<ANTENNA>(.*?)</ANTENNA>', notice_text, re.DOTALL)

    azimuths = []
    gains = []
    heights = []
    powers = []

    for antenna in antenna_blocks:
        azm_match = re.search(r't_azm_max_e\s*=\s*(.+)', antenna)
        gain_match = re.search(r't_gain_max\s*=\s*(.+)', antenna)
        height_match = re.search(r't_hgt_agl\s*=\s*(.+)', antenna)
        power_match = re.search(r't_pwr_ant\s*=\s*(.+)', antenna)

        if azm_match:
            azimuths.append(azm_match.group(1).strip())
        if gain_match:
            gains.append(gain_match.group(1).strip())
        if height_match:
            heights.append(height_match.group(1).strip())
        if power_match:
            powers.append(power_match.group(1).strip())

    # Объединяем через точку
    data['azimuths'] = '.'.join(azimuths)
    data['gains'] = '.'.join(set(gains))  # Уникальные значения
    data['heights'] = '.'.join(set(heights))
    data['powers'] = '.'.join(set(powers))

    return data


def parse_txt_file(file_path, freq_type='tx'):
    """Парсит txt файл и возвращает список данных всех станций
    freq_type: 'tx' для передачи (T12), 'rx' для приема (T13)
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Разделяем на блоки NOTICE
    notice_blocks = re.findall(r'<NOTICE>(.*?)</NOTICE>', content, re.DOTALL)

    stations_data = []
    for notice in notice_blocks:
        data = parse_notice_block(notice)
        data['freq_type'] = freq_type  # Помечаем тип частоты
        stations_data.append(data)

    return stations_data


def merge_tx_rx_data(data_list):
    """Объединяет данные T12 (передача) и T13 (прием) по названию станции"""
    merged = {}

    for data in data_list:
        site_name = data.get('t_site_name', '')
        freq = data.get('t_freq_assgn', '')
        freq_type = data.get('freq_type', 'tx')

        # Создаем уникальный ключ по названию станции
        if site_name not in merged:
            merged[site_name] = data.copy()
            merged[site_name]['freq_tx'] = ''
            merged[site_name]['freq_rx'] = ''

        # Записываем частоту в нужное поле
        if freq_type == 'tx':
            merged[site_name]['freq_tx'] = freq
        else:  # rx
            merged[site_name]['freq_rx'] = freq

    return list(merged.values())


def determine_sheet_from_filename(filename):
    """Определяет лист Excel по названию файла"""
    filename_upper = filename.upper()

    # UZB файлы идут на регистрацию в МСЭ
    if 'UZB' in filename_upper:
        return 'на рег. в МСЭ'
    # Файлы по странам
    elif 'KAZ' in filename_upper:
        return 'КАЗ'
    elif 'TJK' in filename_upper or 'TDK' in filename_upper:
        return 'ТЖК'
    elif 'KGZ' in filename_upper:
        return 'КГЗ'
    elif 'TKM' in filename_upper or 'TUR' in filename_upper:
        return 'ТКМ'
    else:
        return 'на рег. в МСЭ'  # По умолчанию


def create_sheet_with_data(ws, all_data, sheet_type="standard"):
    """Создает лист с данными и форматированием"""

    # Заголовок - первая строка (объединенная)
    ws.merge_cells('A1:Q1')
    header_cell = ws['A1']
    header_cell.value = "Учёт статистических данных по частотоприрсвоениям направленных на координацию с другими Администрациями связи (ИСХ СПС)"
    header_cell.font = Font(bold=True, size=11)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Подзаголовки - третья строка
    if sheet_type == "brific":
        headers = [
            "Название станций\n(Пункт установки)",
            "Координаты\nдолгот",
            "широт",
            "Частота, МГц\nпередача",
            "прием",
            "Ширина\nполосы",
            "Мощн.\nдБВт",
            "КУА,\nдБ",
            "Hант.,\nм",
            "Азимут",
            "№ исходящего письма\n№ письма",
            "дата",
            "Fragment",
            "BRIFIC\nID",
            "Част",
            "Примечание",
            "Исполнитель",
            "ID UZB"
        ]
    else:
        headers = [
            "Название станции\n(пункт установки)",
            "Координаты\nдолгот",
            "широт",
            "Частота, МГц\nпередача",
            "прием",
            "Ширина\nполосы",
            "Мощн.\nдБВт",
            "КУА,\nдБ",
            "Hант.,\nм",
            "Азимут",
            "№ исходящего письма\n№ письма",
            "дата",
            "Ответное письмо\n№ письма",
            "дата",
            "Результат",
            "Направлено в\nБРИФИК",
            "Примечание",
            "Исполнитель",
            "ID UZB"
        ]

    # Стиль заголовка
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Записываем заголовки в третью строку
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    # Записываем данные начиная с 4-й строки
    row = 4
    for data in all_data:
        # Координаты (разделенные)
        long_coord = convert_coordinates(data.get('t_long', ''))
        lat_coord = convert_coordinates(data.get('t_lat', ''))

        ws.cell(row, 1).value = data.get('t_site_name', '')  # Название станции
        ws.cell(row, 2).value = long_coord  # Долгота
        ws.cell(row, 3).value = lat_coord  # Широта
        ws.cell(row, 4).value = data.get('freq_tx', data.get('t_freq_assgn', ''))  # Частота передача
        ws.cell(row, 5).value = data.get('freq_rx', '')  # Частота прием
        ws.cell(row, 6).value = data.get('t_bdwdth_cde', '')  # Ширина
        ws.cell(row, 7).value = data.get('powers', '')  # Мощность
        ws.cell(row, 8).value = data.get('gains', '')  # КУА
        ws.cell(row, 9).value = data.get('heights', '')  # Высота
        ws.cell(row, 10).value = data.get('azimuths', '')  # Азимут
        ws.cell(row, 11).value = ""  # № письма (пустое)
        ws.cell(row, 12).value = convert_date(data.get('t_d_adm_ntc', ''))  # Дата

        if sheet_type == "brific":
            ws.cell(row, 13).value = ""  # Fragment
            ws.cell(row, 14).value = ""  # BRIFIC ID
            ws.cell(row, 15).value = ""  # Част
            ws.cell(row, 16).value = ""  # Примечание
            ws.cell(row, 17).value = ""  # Исполнитель
            ws.cell(row, 18).value = data.get('t_adm_ref_id', '')  # ID UZB
        else:
            ws.cell(row, 13).value = ""  # Ответное письмо №
            ws.cell(row, 14).value = convert_date(data.get('t_d_inuse', ''))  # Дата ввода
            ws.cell(row, 15).value = ""  # Результат
            ws.cell(row, 16).value = ""  # Направлено в БРИФИК
            ws.cell(row, 17).value = ""  # Примечание
            ws.cell(row, 18).value = ""  # Исполнитель
            ws.cell(row, 19).value = data.get('t_adm_ref_id', '')  # ID UZB

        # Применяем стиль к ячейкам
        max_col = 18 if sheet_type == "brific" else 19
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9)

        row += 1

    # Настраиваем ширину столбцов
    column_widths = [18, 10, 10, 10, 10, 10, 9, 9, 8, 15, 20, 12, 15, 12, 12, 15, 20, 15, 15]
    for col, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

    # Высота строк
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 35


def create_excel(data_by_sheet, output_file):
    """Создает Excel файл с несколькими листами"""
    wb = Workbook()

    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Создаем листы в определенном порядке
    sheet_names = ["КГЗ", "ТЖК", "КАЗ", "ТКМ", "на рег. в МСЭ"]

    for sheet_name in sheet_names:
        ws = wb.create_sheet(sheet_name)

        # Определяем тип листа
        sheet_type = "brific" if sheet_name == "на рег. в МСЭ" else "standard"

        # Получаем данные для этого листа
        sheet_data = data_by_sheet.get(sheet_name, [])

        # Создаем лист с данными
        create_sheet_with_data(ws, sheet_data, sheet_type)

    wb.save(output_file)
    print(f"✓ Excel файл создан: {output_file}")


def main():
    """Основная функция"""
    # Папка с txt файлами
    input_folder = input("Введите путь к папке с .txt файлами: ").strip()

    if not os.path.exists(input_folder):
        print("❌ Папка не найдена!")
        return

    # Находим все txt файлы
    txt_files = [f for f in os.listdir(input_folder) if f.endswith('.txt')]

    if not txt_files:
        print("❌ В папке нет .txt файлов!")
        return

    print(f"Найдено {len(txt_files)} файлов\n")

    # Словарь для хранения данных по листам
    data_by_sheet = {
        'КГЗ': [],
        'ТЖК': [],
        'КАЗ': [],
        'ТКМ': [],
        'на рег. в МСЭ': []
    }

    # Группируем файлы по парам T12/T13
    file_groups = {}
    for txt_file in txt_files:
        # Определяем базовое имя (без T12/T13)
        if 'T12' in txt_file.upper():
            base_name = txt_file.upper().replace('T12', 'T1X')
            freq_type = 'tx'
        elif 'T13' in txt_file.upper():
            base_name = txt_file.upper().replace('T13', 'T1X')
            freq_type = 'rx'
        else:
            base_name = txt_file.upper()
            freq_type = 'tx'

        if base_name not in file_groups:
            file_groups[base_name] = {'tx': None, 'rx': None}

        file_groups[base_name][freq_type] = txt_file

    # Обрабатываем группы файлов
    total_stations = 0
    for base_name, files in file_groups.items():
        tx_file = files['tx']
        rx_file = files['rx']

        all_data = []

        # Обрабатываем T12 (передача)
        if tx_file:
            file_path = os.path.join(input_folder, tx_file)
            print(f"Обработка: {tx_file}...")
            stations = parse_txt_file(file_path, 'tx')
            all_data.extend(stations)
            print(f"  └─ Извлечено {len(stations)} станций (передача)")

        # Обрабатываем T13 (прием)
        if rx_file:
            file_path = os.path.join(input_folder, rx_file)
            print(f"Обработка: {rx_file}...")
            stations = parse_txt_file(file_path, 'rx')
            all_data.extend(stations)
            print(f"  └─ Извлечено {len(stations)} станций (прием)")

        if all_data:
            # Объединяем данные T12 и T13
            merged_data = merge_tx_rx_data(all_data)

            # Определяем целевой лист (используем имя любого из файлов)
            sample_file = tx_file if tx_file else rx_file
            target_sheet = determine_sheet_from_filename(sample_file)

            # Добавляем данные на соответствующий лист
            data_by_sheet[target_sheet].extend(merged_data)

            total_stations += len(merged_data)
            print(f"  ✓ Объединено в {len(merged_data)} записей → лист '{target_sheet}'\n")

    print(f"📊 Всего станций: {total_stations}")
    print("\n📋 Распределение по листам:")
    for sheet_name, data in data_by_sheet.items():
        if data:
            print(f"  • {sheet_name}: {len(data)} станций")
        else:
            print(f"  • {sheet_name}: 0 станций (пустой)")

    # Создаем Excel файл с уникальным именем
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(input_folder, f"Учёт_данных_частот_{timestamp}.xlsx")
    create_excel(data_by_sheet, output_file)

    print(f"\n✅ Готово! Данные сохранены в: {output_file}")


if __name__ == "__main__":
    main()