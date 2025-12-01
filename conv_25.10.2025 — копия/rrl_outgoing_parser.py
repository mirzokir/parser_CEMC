import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def convert_coordinates(coord_str):
    """–ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –∏–∑ —Ñ–æ—Ä–º–∞—Ç–∞ +0691949 –≤ 69-19-49"""
    if not coord_str:
        return ""

    coord_str = coord_str.strip('+')
    if len(coord_str) == 7:
        return f"{coord_str[0:2]}-{coord_str[2:4]}-{coord_str[4:7]}"
    elif len(coord_str) == 6:
        return f"{coord_str[0:2]}-{coord_str[2:4]}-{coord_str[4:6]}"
    return coord_str


def parse_head_section(content):
    """–ò–∑–≤–ª–µ–∫–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –∏–∑ —Å–µ–∫—Ü–∏–∏ HEAD"""
    head_data = {}
    head_match = re.search(r'<HEAD>(.*?)</HEAD>', content, re.DOTALL)

    if head_match:
        head_content = head_match.group(1)

        # –ò–∑–≤–ª–µ–∫–∞–µ–º t_adm (—Å—Ç—Ä–∞–Ω–∞)
        adm_match = re.search(r't_adm\s*=\s*(.+)', head_content)
        if adm_match:
            head_data['t_adm'] = adm_match.group(1).strip()

        # –ò–∑–≤–ª–µ–∫–∞–µ–º t_d_sent (–¥–∞—Ç–∞ –æ—Ç–ø—Ä–∞–≤–∫–∏)
        sent_match = re.search(r't_d_sent\s*=\s*(.+)', head_content)
        if sent_match:
            head_data['t_d_sent'] = sent_match.group(1).strip()

    return head_data


def parse_notice_block(notice_text):
    """–ü–∞—Ä—Å–∏—Ç –æ–¥–∏–Ω –±–ª–æ–∫ NOTICE –∏ –∏–∑–≤–ª–µ–∫–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ"""
    data = {}

    # –ò–∑–≤–ª–µ–∫–∞–µ–º –æ—Å–Ω–æ–≤–Ω—ã–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã
    patterns = {
        't_site_name': r't_site_name\s*=\s*(.+)',
        't_freq_assgn': r't_freq_assgn\s*=\s*(.+)',
        't_long': r't_long\s*=\s*(.+)',
        't_lat': r't_lat\s*=\s*(.+)',
        't_bdwdth_cde': r't_bdwdth_cde\s*=\s*(.+)',
        't_d_adm_ntc': r't_d_adm_ntc\s*=\s*(.+)',
        't_adm_ref_id': r't_adm_ref_id\s*=\s*(.+)',
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, notice_text)
        if match:
            data[key] = match.group(1).strip()

    # –ò–∑–≤–ª–µ–∫–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∞–Ω—Ç–µ–Ω–Ω
    antenna_blocks = re.findall(r'<ANTENNA>(.*?)</ANTENNA>', notice_text, re.DOTALL)

    for antenna in antenna_blocks:
        gain_match = re.search(r't_gain_max\s*=\s*(.+)', antenna)
        height_match = re.search(r't_hgt_agl\s*=\s*(.+)', antenna)
        power_match = re.search(r't_pwr_dbw\s*=\s*(.+)', antenna)

        if gain_match:
            data['t_gain_max'] = gain_match.group(1).strip()
        if height_match:
            data['t_hgt_agl'] = height_match.group(1).strip()
        if power_match:
            data['t_pwr_dbw'] = power_match.group(1).strip()

        # –ò–∑–≤–ª–µ–∫–∞–µ–º –¥–∞–Ω–Ω—ã–µ –ø—Ä–∏–Ω–∏–º–∞—é—â–µ–π —Å—Ç–∞–Ω—Ü–∏–∏
        rx_match = re.search(r'<RX_STATION>(.*?)</RX_STATION>', antenna, re.DOTALL)
        if rx_match:
            rx_content = rx_match.group(1)
            rx_site_match = re.search(r't_site_name\s*=\s*(.+)', rx_content)
            if rx_site_match:
                data['rx_site_name'] = rx_site_match.group(1).strip()

    return data


def parse_txt_file(file_path):
    """–ü–∞—Ä—Å–∏—Ç txt —Ñ–∞–π–ª –∏ –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å–ø–∏—Å–æ–∫ –¥–∞–Ω–Ω—ã—Ö –≤—Å–µ—Ö —Å—Ç–∞–Ω—Ü–∏–π"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # –ò–∑–≤–ª–µ–∫–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ HEAD
    head_data = parse_head_section(content)

    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —ç—Ç–æ —Ñ–∞–π–ª UZB (–∏—Å—Ö–æ–¥—è—â–∏–µ)
    t_adm = head_data.get('t_adm', '').upper()
    if 'UZB' not in t_adm:
        return [], head_data, False

    # –†–∞–∑–¥–µ–ª—è–µ–º –Ω–∞ –±–ª–æ–∫–∏ NOTICE
    notice_blocks = re.findall(r'<NOTICE>(.*?)</NOTICE>', content, re.DOTALL)

    stations_data = []
    for notice in notice_blocks:
        data = parse_notice_block(notice)
        # –î–æ–±–∞–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ HEAD
        data['t_adm'] = head_data.get('t_adm', '')
        data['t_d_sent'] = head_data.get('t_d_sent', '')
        stations_data.append(data)

    return stations_data, head_data, True


def link_stations(stations_data):
    """–°–≤—è–∑—ã–≤–∞–µ—Ç —Å—Ç–∞–Ω—Ü–∏–∏ –∏ –æ–ø—Ä–µ–¥–µ–ª—è–µ—Ç —á–∞—Å—Ç–æ—Ç—ã –ø—Ä–∏—ë–º–∞"""
    # –°–æ–∑–¥–∞—ë–º —Å–ª–æ–≤–∞—Ä—å –¥–ª—è –±—ã—Å—Ç—Ä–æ–≥–æ –ø–æ–∏—Å–∫–∞ —Å—Ç–∞–Ω—Ü–∏–π –ø–æ –∏–º–µ–Ω–∏
    station_map = {}
    for station in stations_data:
        site_name = station.get('t_site_name', '')
        station_map[site_name] = station

    # –î–ª—è –∫–∞–∂–¥–æ–π —Å—Ç–∞–Ω—Ü–∏–∏ –Ω–∞—Ö–æ–¥–∏–º —á–∞—Å—Ç–æ—Ç—É –ø—Ä–∏—ë–º–∞
    for station in stations_data:
        rx_site_name = station.get('rx_site_name', '')
        if rx_site_name and rx_site_name in station_map:
            # –ß–∞—Å—Ç–æ—Ç–∞ –ø—Ä–∏—ë–º–∞ = —á–∞—Å—Ç–æ—Ç–∞ –ø–µ—Ä–µ–¥–∞—á–∏ –ø–∞—Ä–Ω–æ–π —Å—Ç–∞–Ω—Ü–∏–∏
            rx_station = station_map[rx_site_name]
            station['freq_rx'] = rx_station.get('t_freq_assgn', '')
        else:
            station['freq_rx'] = ''

    return stations_data


def create_sheet_with_data(ws, all_data):
    """–°–æ–∑–¥–∞–µ—Ç –ª–∏—Å—Ç —Å –¥–∞–Ω–Ω—ã–º–∏ –¥–ª—è –ò–°–•–û–î–Ø–©–ò–ï –†–†–õ"""

    # –ó–∞–≥–æ–ª–æ–≤–æ–∫ - –ø–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ (–æ–±—ä–µ–¥–∏–Ω–µ–Ω–Ω–∞—è)
    ws.merge_cells('A1:Q1')
    header_cell = ws['A1']
    header_cell.value = "–£—á—ë—Ç —Å—Ç–∞—Ç–∏—Å—Ç–∏—á–µ—Å–∫–∏—Ö –¥–∞–Ω–Ω—ã—Ö –ø–æ —á–∞—Å—Ç–æ—Ç–æ–ø—Ä–∏—Ä—Å–≤–æ–µ–Ω–∏—è–º –Ω–∞–ø—Ä–∞–≤–ª–µ–Ω–Ω—ã—Ö –Ω–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ü–∏—é —Å –ê–° –†–£–∑ (–ò–°–•–û–î–Ø–©–ò–ï)-–†–†–õ"
    header_cell.font = Font(bold=True, size=11)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # –í—Ç–æ—Ä–∞—è —Å—Ç—Ä–æ–∫–∞ - –æ–±—ä–µ–¥–∏–Ω–µ–Ω–Ω—ã–µ –∑–∞–≥–æ–ª–æ–≤–∫–∏
    ws.merge_cells('A2:B2')
    ws['A2'].value = "–ß–∞—Å—Ç–æ—Ç–∞, –ú–ì—Ü"

    ws.merge_cells('C2:D2')
    ws['C2'].value = "–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã"

    ws.merge_cells('J2:K2')
    ws['J2'].value = "‚Ññ –∏ –¥–∞—Ç–∞ –≤—Ö–æ–¥—è—â–µ–≥–æ –ø–∏—Å—å–º–∞"

    ws.merge_cells('L2:M2')
    ws['L2'].value = "‚Ññ –∏ –¥–∞—Ç–∞ –∏—Å—Ö–æ–¥—è—â–µ–≥–æ –ø–∏—Å—å–º–∞"

    # –¢—Ä–µ—Ç—å—è —Å—Ç—Ä–æ–∫–∞ - –ø–æ–¥–∑–∞–≥–æ–ª–æ–≤–∫–∏
    headers_row3 = [
        "–ø–µ—Ä–µ–¥–∞—á–∞", "–ø—Ä–∏—ë–º",
        "–¥–æ–ª–≥–æ—Ç–∞", "—à–∏—Ä–æ—Ç–∞",
        "–ü—É–Ω–∫—Ç —É—Å—Ç–∞–Ω–æ–≤–∫–∏",
        "–®–∏—Ä–∏–Ω–∞\n–ø–æ–ª–æ—Å—ã,\n–ú–ì—Ü",
        "–ö–æ—ç—Ñ-—Ç\n—É—Å–∏–ª–µ–Ω–∏—è,\n–¥–ë",
        "–ú–æ—â–Ω–æ—Å—Ç—å\n–ø–µ—Ä–µ–¥–∞—Ç—á–∏–∫–∞,\n–¥–ë–í—Ç",
        "–í—ã—Å–æ—Ç–∞\n–∞–Ω—Ç–µ–Ω–Ω—ã, –º",
        "–ø–µ—Ä–≤–∏—á–Ω–æ–µ", "–ø–æ–≤—Ç–æ—Ä–Ω–æ–µ",
        "–ø–µ—Ä–≤–∏—á–Ω–æ–µ", "–ø–æ–≤—Ç–æ—Ä–Ω–æ–µ",
        "–†–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ–≥–ª–∞—Å–æ–≤–∞–Ω–∏—è\n(—Å–æ–≥–ª–∞—Å–æ–≤–∞–Ω–æ/\n–Ω–µ —Å–æ–≥–ª–∞—Å–æ–≤–∞–Ω–æ)",
        "–ü—Ä–∏–º–µ—á–∞–Ω–∏–µ",
        "–ò—Å–ø–æ–ª–Ω–∏—Ç–µ–ª—å",
        "id1/ unique id given by\nthe administration to\nthe assignment"
    ]

    # –°—Ç–∏–ª—å –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # –ü—Ä–∏–º–µ–Ω—è–µ–º —Å—Ç–∏–ª–∏ –∫ —Å—Ç—Ä–æ–∫–µ 2
    for col in range(1, 18):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –ø–æ–¥–∑–∞–≥–æ–ª–æ–≤–∫–∏ –≤ —Ç—Ä–µ—Ç—å—é —Å—Ç—Ä–æ–∫—É
    for col, header in enumerate(headers_row3, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ –Ω–∞—á–∏–Ω–∞—è —Å 4-–π —Å—Ç—Ä–æ–∫–∏
    row = 4
    for data in all_data:
        # –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã
        long_coord = convert_coordinates(data.get('t_long', ''))
        lat_coord = convert_coordinates(data.get('t_lat', ''))

        ws.cell(row, 1).value = data.get('t_freq_assgn', '')  # –ß–∞—Å—Ç–æ—Ç–∞ –ø–µ—Ä–µ–¥–∞—á–∞
        ws.cell(row, 2).value = data.get('freq_rx', '')  # –ß–∞—Å—Ç–æ—Ç–∞ –ø—Ä–∏—ë–º
        ws.cell(row, 3).value = long_coord  # –î–æ–ª–≥–æ—Ç–∞
        ws.cell(row, 4).value = lat_coord  # –®–∏—Ä–æ—Ç–∞
        ws.cell(row, 5).value = data.get('t_site_name', '')  # –ü—É–Ω–∫—Ç —É—Å—Ç–∞–Ω–æ–≤–∫–∏
        ws.cell(row, 6).value = data.get('t_bdwdth_cde', '')  # –®–∏—Ä–∏–Ω–∞ –ø–æ–ª–æ—Å—ã
        ws.cell(row, 7).value = data.get('t_gain_max', '')  # –ö–æ—ç—Ñ —É—Å–∏–ª–µ–Ω–∏—è
        ws.cell(row, 8).value = data.get('t_pwr_dbw', '')  # –ú–æ—â–Ω–æ—Å—Ç—å
        ws.cell(row, 9).value = data.get('t_hgt_agl', '')  # –í—ã—Å–æ—Ç–∞

        # –§–æ—Ä–º–∏—Ä—É–µ–º –Ω–æ–º–µ—Ä –≤—Ö–æ–¥—è—â–µ–≥–æ: t_d_sent + t_d_adm_ntc
        incoming_number = ""
        d_sent = data.get('t_d_sent', '')
        d_adm_ntc = data.get('t_d_adm_ntc', '')
        if d_sent and d_adm_ntc:
            incoming_number = f"{d_sent}/{d_adm_ntc}"
        elif d_sent:
            incoming_number = d_sent
        elif d_adm_ntc:
            incoming_number = d_adm_ntc

        ws.cell(row, 10).value = incoming_number  # ‚Ññ –≤—Ö–æ–¥—è—â–µ–≥–æ –ø–µ—Ä–≤–∏—á–Ω–æ–µ
        ws.cell(row, 11).value = ""  # ‚Ññ –≤—Ö–æ–¥—è—â–µ–≥–æ –ø–æ–≤—Ç–æ—Ä–Ω–æ–µ
        ws.cell(row, 12).value = ""  # ‚Ññ –∏—Å—Ö–æ–¥—è—â–µ–≥–æ –ø–µ—Ä–≤–∏—á–Ω–æ–µ
        ws.cell(row, 13).value = ""  # ‚Ññ –∏—Å—Ö–æ–¥—è—â–µ–≥–æ –ø–æ–≤—Ç–æ—Ä–Ω–æ–µ
        ws.cell(row, 14).value = ""  # –†–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ–≥–ª–∞—Å–æ–≤–∞–Ω–∏—è
        ws.cell(row, 15).value = ""  # –ü—Ä–∏–º–µ—á–∞–Ω–∏–µ
        ws.cell(row, 16).value = ""  # –ò—Å–ø–æ–ª–Ω–∏—Ç–µ–ª—å
        ws.cell(row, 17).value = data.get('t_adm_ref_id', '')  # id1

        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Å—Ç–∏–ª—å –∫ —è—á–µ–π–∫–∞–º
        for col in range(1, 18):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9)

        row += 1

    # –ù–∞—Å—Ç—Ä–∞–∏–≤–∞–µ–º —à–∏—Ä–∏–Ω—É —Å—Ç–æ–ª–±—Ü–æ–≤
    column_widths = [12, 12, 10, 10, 20, 10, 10, 12, 10, 15, 15, 15, 15, 15, 15, 15, 25]
    for col, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

    # –í—ã—Å–æ—Ç–∞ —Å—Ç—Ä–æ–∫
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 50


def create_excel(all_data, output_file):
    """–°–æ–∑–¥–∞–µ—Ç Excel —Ñ–∞–π–ª —Å –æ–¥–Ω–∏–º –ª–∏—Å—Ç–æ–º"""
    wb = Workbook()

    # –£–¥–∞–ª—è–µ–º –¥–µ—Ñ–æ–ª—Ç–Ω—ã–π –ª–∏—Å—Ç
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # –°–æ–∑–¥–∞–µ–º –µ–¥–∏–Ω—Å—Ç–≤–µ–Ω–Ω—ã–π –ª–∏—Å—Ç
    ws = wb.create_sheet("–ò–°–•–û–î–Ø–©–ò–ï –†–†–õ")

    # –°–æ–∑–¥–∞–µ–º –ª–∏—Å—Ç —Å –¥–∞–Ω–Ω—ã–º–∏
    create_sheet_with_data(ws, all_data)

    wb.save(output_file)
    print(f"‚úì Excel —Ñ–∞–π–ª —Å–æ–∑–¥–∞–Ω: {output_file}")


def main():
    """–û—Å–Ω–æ–≤–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è"""
    # –ü–∞–ø–∫–∞ —Å txt —Ñ–∞–π–ª–∞–º–∏
    input_folder = input("–í–≤–µ–¥–∏—Ç–µ –ø—É—Ç—å –∫ –ø–∞–ø–∫–µ —Å .txt —Ñ–∞–π–ª–∞–º–∏ (–ò–°–•–û–î–Ø–©–ò–ï –†–†–õ - UZB): ").strip()

    if not os.path.exists(input_folder):
        print("‚ùå –ü–∞–ø–∫–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞!")
        return

    # –ù–∞—Ö–æ–¥–∏–º –≤—Å–µ txt —Ñ–∞–π–ª—ã
    txt_files = [f for f in os.listdir(input_folder) if f.endswith('.txt')]

    if not txt_files:
        print("‚ùå –í –ø–∞–ø–∫–µ –Ω–µ—Ç .txt —Ñ–∞–π–ª–æ–≤!")
        return

    print(f"–ù–∞–π–¥–µ–Ω–æ {len(txt_files)} —Ñ–∞–π–ª–æ–≤\n")

    # –°–ø–∏—Å–æ–∫ –¥–ª—è –≤—Å–µ—Ö –¥–∞–Ω–Ω—ã—Ö
    all_data = []

    # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª—ã
    uzb_files_count = 0
    for txt_file in txt_files:
        file_path = os.path.join(input_folder, txt_file)
        print(f"–û–±—Ä–∞–±–æ—Ç–∫–∞: {txt_file}...")

        stations_data, head_data, is_uzb = parse_txt_file(file_path)

        if not is_uzb:
            print(f"  ‚ö†Ô∏è  –ü—Ä–æ–ø—É—Å–∫–∞–µ–º (–Ω–µ UZB —Ñ–∞–π–ª: {head_data.get('t_adm', 'N/A')})\n")
            continue

        uzb_files_count += 1

        # –°–≤—è–∑—ã–≤–∞–µ–º —Å—Ç–∞–Ω—Ü–∏–∏ –∏ –æ–ø—Ä–µ–¥–µ–ª—è–µ–º —á–∞—Å—Ç–æ—Ç—ã –ø—Ä–∏—ë–º–∞
        stations_data = link_stations(stations_data)

        # –î–æ–±–∞–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ
        all_data.extend(stations_data)

        print(f"  ‚îî‚îÄ –ò–∑–≤–ª–µ—á–µ–Ω–æ {len(stations_data)} —Å—Ç–∞–Ω—Ü–∏–π –æ—Ç UZB\n")

    if uzb_files_count == 0:
        print("‚ùå –ù–µ –Ω–∞–π–¥–µ–Ω–æ —Ñ–∞–π–ª–æ–≤ —Å t_adm=UZB!")
        return

    print(f"üìä –í—Å–µ–≥–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ UZB —Ñ–∞–π–ª–æ–≤: {uzb_files_count}")
    print(f"üìä –í—Å–µ–≥–æ —Å—Ç–∞–Ω—Ü–∏–π: {len(all_data)}")

    # –°–æ–∑–¥–∞–µ–º Excel —Ñ–∞–π–ª —Å —É–Ω–∏–∫–∞–ª—å–Ω—ã–º –∏–º–µ–Ω–µ–º
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(input_folder, f"–ò–°–•–û–î–Ø–©–ò–ï_–†–†–õ_{timestamp}.xlsx")
    create_excel(all_data, output_file)

    print(f"\n‚úÖ –ì–æ—Ç–æ–≤–æ! –î–∞–Ω–Ω—ã–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤: {output_file}")


if __name__ == "__main__":
    main()